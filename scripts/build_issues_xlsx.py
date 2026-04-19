from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = Workbook()
ws = wb.active
ws.title = "Issues Tracker"

# ---- Data ----
# Columns: Sno, Title, Issue Description, Severity, Assignee, Start Date, End Date, Status
# Dates laid out against the prioritized roadmap. Today is 2026-04-18 (Sat).
# Week 1: 2026-04-20 -> 2026-04-24
# Week 2: 2026-04-27 -> 2026-05-01
# Month 2 (ideal): 2026-05-04 -> 2026-05-22
# Background/ongoing: flexible windows later

issues = [
    # --- WEEK 1 (Critical) ---
    (1, "Webhook HMAC signatures for Edge Functions",
     "screenshot-webhook and activity-webhook have verify_jwt=false and no HMAC. Any actor with the URL can POST into any tenant's data. Add HMAC-SHA256 signature + timestamp header and verify with crypto.timingSafeEqual. Per-org secret stored in organizations.webhook_secret.",
     "Critical", "Security Engineer", date(2026, 4, 20), date(2026, 4, 20), "Not Started"),

    (2, "Enforce table allowlist in Forge proxy",
     "forge-proxy-controller.js only logs warnings on sensitive tables. Service-role client bypasses RLS. Convert SENSITIVE_TABLES to a hard PROXY_POLICY allowlist with per-table ops, required filters, and column allowlists. Default-deny.",
     "Critical", "Backend Engineer", date(2026, 4, 20), date(2026, 4, 20), "Not Started"),

    (3, "Add Postgres advisory lock around background services",
     "activity-polling (setInterval 3min), clustering (setTimeout daily), notifications services are single-instance with no distributed lock. Under Cloud Run scale-out they duplicate work and load. Wrap each tick in pg_try_advisory_lock(hashtext(serviceName)).",
     "Critical", "Backend Engineer", date(2026, 4, 21), date(2026, 4, 21), "Not Started"),

    (4, "Move admin dashboard sessions to Supabase table",
     "admin-dashboard-controller.js stores sessions in a Map. Cloud Run horizontal scaling / deploys wipe sessions. Create admin_sessions table, wrap set/get/delete in an adminSessionStore module, keep 30s in-process LRU cache for latency.",
     "Critical", "Backend Engineer", date(2026, 4, 22), date(2026, 4, 22), "Not Started"),

    (5, "Remove asApp() fallback in scheduled worklog sync",
     "scheduledWorklogSync.js falls back to createJiraWorklogAsApp when asUser throws AUTH_TYPE_UNAVAILABLE, breaking per-user attribution (the product's core value). Write a PENDING row instead so interactive sync creates the worklog with correct user attribution on next app open.",
     "Critical", "Forge/Backend Engineer", date(2026, 4, 23), date(2026, 4, 23), "Not Started"),

    # --- WEEK 2 (Important) ---
    (6, "Delete update-issues-cache stub Edge Function",
     "supabase/functions/update-issues-cache/index.ts is an explicit placeholder (TODO: Implement actual cache update mechanism) that returns success without doing anything. Delete the function and any cron calling it. Forge manifest already handles cache refresh via avi:jira:updated:issue.",
     "Important", "Backend Engineer", date(2026, 4, 27), date(2026, 4, 27), "Not Started"),

    (7, "Update Forge KVS clearSiteCache to enumerate keys",
     "lifecycleService.js comment says KVS doesn't support enumeration. Outdated: @forge/kvs now supports kvs.query().where('key', WhereConditions.beginsWith(prefix)).getMany() with cursor pagination. Replace hardcoded two-key delete with query-based cleanup on uninstall.",
     "Important", "Forge Engineer", date(2026, 4, 28), date(2026, 4, 28), "Not Started"),

    (8, "Add 5-minute LRU cache for Atlassian /me endpoint",
     "atlassian-auth.js and dashboard-auth.js hit Atlassian APIs on every request (1-3 calls, ~400-900ms TTFB). Add in-process LRU keyed by sha256(token) with 5-min TTL. For dashboard-auth, cache all three results (/me, /accessible-resources, /mypermissions) together.",
     "Important", "Backend Engineer", date(2026, 4, 28), date(2026, 4, 28), "Not Started"),

    (9, "Add exponential backoff to activity polling service",
     "activity-polling-service.js has fixed 3-min interval with no backoff on failure. Track consecutive failures and multiply interval by 2x up to 30-min cap. Convert from setInterval to self-rescheduling setTimeout.",
     "Important", "Backend Engineer", date(2026, 4, 29), date(2026, 4, 29), "Not Started"),

    (10, "Add size cap and auto-VACUUM for offline SQLite",
     "OfflineManager (desktop_app.py line 2229) has no file size check, no TTL on synced rows. On startup, if DB > 500MB, delete synced rows older than 14 days and run VACUUM INTO. Prevents unbounded growth for offline or sync-blocked users.",
     "Important", "Desktop Engineer", date(2026, 4, 29), date(2026, 4, 29), "Not Started"),

    (11, "Log AI confidence score and add UI transparency tooltip",
     "activity-db-service.js uses MIN_CONFIDENCE_THRESHOLD (default 0.5) to decide assignment silently. Log confidenceScore into activity_records.metadata. Add 'Why unassigned?' tooltip in UI explaining AI confidence was below threshold.",
     "Important", "Full-Stack Engineer", date(2026, 4, 30), date(2026, 4, 30), "Not Started"),

    (12, "Tighten duplicate-user heuristic to equality-only match",
     "issueQueryService.js uses substring match on Atlassian account IDs (otherId.includes(currentId)) which can match unrelated accounts. Remove substring branch; keep email equality match only. Prepare for schema-level fix.",
     "Important", "Forge Engineer", date(2026, 4, 30), date(2026, 4, 30), "Not Started"),

    (13, "Add admin email allowlist alongside password",
     "admin-dashboard-controller.js uses a single shared ADMIN_DASHBOARD_PASSWORD. Add ADMIN_EMAIL_ALLOWLIST env var and require email + password. Stops any single env-var leak from compromising all admin access.",
     "Important", "Backend Engineer", date(2026, 5, 1), date(2026, 5, 1), "Not Started"),

    # --- MONTH 2 (Ideal Architecture) ---
    (14, "Replace polling services with pg-boss on Cloud Run Jobs",
     "Replace three self-scheduled services (activity, clustering, notifications) with pg-boss queues. Use Cloud Scheduler + Cloud Run Jobs for scheduled clustering. Provides exactly-once delivery, retries, dead-letter, and decouples worker from API scaling.",
     "Critical", "Backend / DevOps Engineer", date(2026, 5, 4), date(2026, 5, 8), "Not Started"),

    (15, "Migrate server logs to Pino with structured redaction",
     "Replace regex-based log-sanitizer with Pino redact config keyed on JSON paths (e.g. *.email, *.token). Keep existing regex as second-pass belt-and-suspenders. Emits structured JSON logs queryable in Cloud Logging.",
     "Important", "Backend Engineer", date(2026, 5, 11), date(2026, 5, 13), "Not Started"),

    (16, "Replace generic Forge proxy with named RPC endpoints",
     "Replace POST /api/forge/proxy (table/operation pattern) with named endpoints like /api/forge/activity/query, /api/forge/worklog/sync. Move business logic server-side; make API enumerable. Use a role-scoped Postgres user, not service role.",
     "Critical", "Backend Engineer", date(2026, 5, 11), date(2026, 5, 15), "Not Started"),

    (17, "Migrate admin dashboard to Supabase Auth with MFA + audit log",
     "Replace hand-rolled session tokens with Supabase Auth. Enforce admin allowlist via RLS. Enable TOTP MFA. Write every admin action to an append-only admin_audit table.",
     "Important", "Backend Engineer", date(2026, 5, 18), date(2026, 5, 20), "Not Started"),

    (18, "Fix duplicate users/orgs in schema + one-time data migration",
     "Merge duplicate organizations by jira_cloud_id into oldest row; merge duplicate users by (organization_id, atlassian_account_id). Add unique constraints. Delete the heuristic dual-read code in issueQueryService and scheduledWorklogSync once clean.",
     "Important", "Backend / Data Engineer", date(2026, 5, 18), date(2026, 5, 22), "Not Started"),

    # --- ONGOING / BACKGROUND ---
    (19, "Document CORS no-origin allow as deliberate design choice",
     "ai-server/src/index.js allows no-origin requests for desktop/Forge/server-to-server clients. This is correct but not a security boundary. Expand existing comment to enumerate the three callers and clarify auth is enforced in middleware.",
     "Low", "Backend Engineer", date(2026, 5, 25), date(2026, 5, 25), "Not Started"),

    (20, "Document desktop secure-storage threat model",
     "secure_storage.py already uses OWASP-compliant PBKDF2+AES fallback. Add python-desktop-app/SECURITY.md clarifying threat model: defends against cold-disk extraction and cross-machine copying; does not defend against malware running as the user.",
     "Low", "Desktop / Security Engineer", date(2026, 5, 25), date(2026, 5, 25), "Not Started"),

    (21, "Secure /reset-upload-lock endpoint with admin auth + rate limit",
     "Require admin auth, add 1/min rate limit per user, log every call to admin audit log with caller and target upload ID.",
     "Low", "Backend Engineer", date(2026, 5, 26), date(2026, 5, 26), "Not Started"),

    (22, "Establish quarterly log-sanitizer pattern review",
     "Add to engineering runbook: quarterly grep diff of logger.info|warn|error call sites vs. sanitizer patterns in log-sanitizer.js. Catches drift when new sensitive fields are added.",
     "Low", "Backend Engineer", date(2026, 5, 27), date(2026, 5, 27), "Not Started"),

    (23, "Incrementally split desktop_app.py into modules",
     "563 KB single-file Python app hurts onboarding and test coverage. Split opportunistically by responsibility (tray/, capture/, sync/, idle/) as areas are touched. No big-bang rewrite.",
     "Low", "Desktop Engineer", date(2026, 6, 1), date(2026, 6, 30), "Not Started"),
]

headers = ["Sno", "Title", "Issue Description", "Severity", "Assignee",
           "Start Date", "End Date", "Status"]

# ---- Styling ----
BASE_FONT = "Arial"
header_font = Font(name=BASE_FONT, bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="1F4E78")
body_font = Font(name=BASE_FONT, size=10)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(border_style="thin", color="B4B4B4")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

sev_fills = {
    "Critical": PatternFill("solid", start_color="F8CBAD"),
    "Important": PatternFill("solid", start_color="FFE699"),
    "Low": PatternFill("solid", start_color="C6E0B4"),
}
status_fills = {
    "Not Started": PatternFill("solid", start_color="D9D9D9"),
    "In Progress": PatternFill("solid", start_color="FFF2CC"),
    "Done": PatternFill("solid", start_color="C6E0B4"),
    "Blocked": PatternFill("solid", start_color="F4B084"),
}

# Header row
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# Data rows
for row_idx, row in enumerate(issues, start=2):
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = body_font
        cell.border = border
        # alignment
        if col_idx in (1, 4, 6, 7, 8):
            cell.alignment = center
        else:
            cell.alignment = wrap
        # severity / status colouring
        if col_idx == 4 and value in sev_fills:
            cell.fill = sev_fills[value]
            cell.font = Font(name=BASE_FONT, size=10, bold=True)
        if col_idx == 8 and value in status_fills:
            cell.fill = status_fills[value]
        # date formatting
        if col_idx in (6, 7):
            cell.number_format = "yyyy-mm-dd"

# Column widths
widths = {
    "A": 6,    # Sno
    "B": 42,   # Title
    "C": 80,   # Description
    "D": 12,   # Severity
    "E": 26,   # Assignee
    "F": 13,   # Start
    "G": 13,   # End
    "H": 14,   # Status
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Row heights - let Excel auto-size via wrap + set header
ws.row_dimensions[1].height = 28
for r in range(2, len(issues) + 2):
    ws.row_dimensions[r].height = 70

# Freeze header
ws.freeze_panes = "A2"

# AutoFilter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(issues)+1}"

# ---- Summary sheet ----
ws2 = wb.create_sheet("Summary")

summary_rows = [
    ["BRD Time Tracker — Remediation Issue Tracker", ""],
    ["Generated", "2026-04-18"],
    ["Source", "RISK_REMEDIATION.md"],
    ["", ""],
    ["Totals", ""],
    ["Total issues", f"=COUNTA('Issues Tracker'!A2:A{len(issues)+1})"],
    ["Critical", f'=COUNTIF(\'Issues Tracker\'!D2:D{len(issues)+1},"Critical")'],
    ["Important", f'=COUNTIF(\'Issues Tracker\'!D2:D{len(issues)+1},"Important")'],
    ["Low", f'=COUNTIF(\'Issues Tracker\'!D2:D{len(issues)+1},"Low")'],
    ["", ""],
    ["Status breakdown", ""],
    ["Not Started", f'=COUNTIF(\'Issues Tracker\'!H2:H{len(issues)+1},"Not Started")'],
    ["In Progress", f'=COUNTIF(\'Issues Tracker\'!H2:H{len(issues)+1},"In Progress")'],
    ["Done", f'=COUNTIF(\'Issues Tracker\'!H2:H{len(issues)+1},"Done")'],
    ["Blocked", f'=COUNTIF(\'Issues Tracker\'!H2:H{len(issues)+1},"Blocked")'],
    ["", ""],
    ["Roadmap windows", ""],
    ["Week 1 (Critical fixes)", "2026-04-20 to 2026-04-24"],
    ["Week 2 (Important fixes)", "2026-04-27 to 2026-05-01"],
    ["Month 2 (Ideal architecture)", "2026-05-04 to 2026-05-22"],
    ["Ongoing / background", "2026-05-25 onward"],
]

for r, (k, v) in enumerate(summary_rows, start=1):
    a = ws2.cell(row=r, column=1, value=k)
    b = ws2.cell(row=r, column=2, value=v)
    a.font = Font(name=BASE_FONT, size=11, bold=(r == 1 or k in (
        "Totals", "Status breakdown", "Roadmap windows")))
    b.font = Font(name=BASE_FONT, size=11)
    a.alignment = Alignment(vertical="center")
    b.alignment = Alignment(vertical="center")

ws2["A1"].font = Font(name=BASE_FONT, size=14, bold=True, color="1F4E78")
ws2.column_dimensions["A"].width = 36
ws2.column_dimensions["B"].width = 36

# Move summary to be first tab
wb.move_sheet("Summary", offset=-1)

out_path = "/sessions/lucid-clever-goodall/mnt/JIRAForge/BRD_Remediation_Issues.xlsx"
wb.save(out_path)
print(f"Wrote: {out_path}")
print(f"Issues: {len(issues)}")
